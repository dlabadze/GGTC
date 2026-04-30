from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import openpyxl


class UpdateCategoriesWizard(models.TransientModel):
    _name = 'update.categories.wizard'
    _description = 'Update Categories Wizard'

    categories = fields.Selection(
        selection=[
            ('1', 'Category 1'),
            ('2', 'Category 2'),
            ('3', 'Category 3'),
        ],
        string='Categories',
        required=True,
    )
    
    excel_file = fields.Binary(
        string='Excel File',
        required=True,
        help='Upload an Excel file (.xlsx or .xls)',
    )
    
    file_name = fields.Char(
        string='File Name',
    )
    
    def action_update_categories(self):
        if self.categories == '1':
            self.action_update_categories_1()
        elif self.categories == '2':
            self.action_update_categories_2()
        elif self.categories == '3':
            self.action_update_categories_3()
    
    def action_update_categories_1(self):
        """Read Excel from row 2
        Column C: for searching analytic account (code)
        Column E: for category name
        """
        self.ensure_one()
        
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file.'))
        
        try:
            # Decode the Excel file
            file_data = base64.b64decode(self.excel_file)
            workbook = openpyxl.load_workbook(io.BytesIO(file_data))
            sheet = workbook.active
            
            created_count = 0
            skipped_count = 0
            
            # Iterate through rows (starting from row 2 to skip header)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Column C is index 2 (0-based indexing) - for analytic account
                column_c = row[2] if len(row) > 2 else None
                # Column E is index 4 (0-based indexing) - for category name
                column_e = row[4] if len(row) > 4 else None
                
               
                
                # Check if category already exists
                existing_category = self.env['product.category'].search([
                    ('name', '=', column_e)
                ], limit=1)
                
                if existing_category:
                    skipped_count += 1
                    continue
                
                # Find account.analytic.account by code (column C)
                analytic_account = None
                if column_c:
                    analytic_account = self.env['account.analytic.account'].search([
                        ('code', '=', column_c)
                    ], limit=1)
                
                # Create new category with name from column E
                self.env['product.category'].create({
                    'name': column_e,
                    'x_studio_many2one_field_2o6_1j1dfj1v3': analytic_account.id if analytic_account else None,
                    'property_valuation': 'real_time',
                    'property_cost_method': 'average',
                })
                created_count += 1
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Success'),
                    'message': _('Created %s categories, skipped %s existing') % (created_count, skipped_count),
                    'type': 'success',
                    'sticky': False,
                }
            }
            
        except Exception as e:
            raise UserError(_('Error processing Excel file: %s') % str(e))

    def action_update_categories_2(self):
        """Read Excel from row 2
        Column C: for searching analytic account (code)
        Column E: parent category name
        Column F: category name
        """
        self.ensure_one()
        
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file.'))
        
        try:
            # Decode the Excel file
            file_data = base64.b64decode(self.excel_file)
            workbook = openpyxl.load_workbook(io.BytesIO(file_data))
            sheet = workbook.active
            
            created_count = 0
            skipped_count = 0
            
            # Iterate through rows (starting from row 2 to skip header)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Column C is index 2 - for analytic account
                column_c = row[2] if len(row) > 2 else None
                # Column E is index 4 - for parent category name
                column_e = row[4] if len(row) > 4 else None
                # Column F is index 5 - for category name
                column_f = row[5] if len(row) > 5 else None
                
                # Skip if column F is empty
                if not column_f:
                    continue
                
                # Check if category already exists
                parent_category = None
                existing_category = None
                if column_e:
                    parent_category = self.env['product.category'].search([
                        ('name', '=', column_e)
                    ], limit=1)
                
                    existing_category = self.env['product.category'].search([
                        ('name', '=', column_f),
                        ('parent_id', '=', parent_category.id),
                    ], limit=1)
                else:
                    existing_category = self.env['product.category'].search([
                        ('name', '=', column_f)
                    ], limit=1)

            
                if existing_category:
                    skipped_count += 1
                    continue
                
                # Find parent category (column E) if exists
               
                
                # Find account.analytic.account by code (column C)
                analytic_account = None
                if column_c:
                    analytic_account = self.env['account.analytic.account'].search([
                        ('code', '=', column_c)
                    ], limit=1)
                
                # Create new category with name from column F
                self.env['product.category'].create({
                    'name': column_f,
                    'parent_id': parent_category.id if parent_category else None,
                    'x_studio_many2one_field_2o6_1j1dfj1v3': analytic_account.id if analytic_account else None,
                    'property_valuation': 'real_time',
                    'property_cost_method': 'average',
                })
                created_count += 1
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Success'),
                    'message': _('Created %s categories, skipped %s existing') % (created_count, skipped_count),
                    'type': 'success',
                    'sticky': False,
                }
            }
            
        except Exception as e:
            raise UserError(_('Error processing Excel file: %s') % str(e))

    def action_update_categories_3(self):
        """Read Excel from row 2
        Column C: for searching analytic account (code)
        Column E: grandparent category name (parent.parent)
        Column F: parent category name
        Column G: category name
        Hierarchy: E -> F -> G
        """
        self.ensure_one()
        
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file.'))
        
        try:
            # Decode the Excel file
            file_data = base64.b64decode(self.excel_file)
            workbook = openpyxl.load_workbook(io.BytesIO(file_data))
            sheet = workbook.active
            
            created_count = 0
            skipped_count = 0
            
            # Iterate through rows (starting from row 2 to skip header)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Column C is index 2 - for analytic account
                column_c = row[2] if len(row) > 2 else None
                # Column E is index 4 - for grandparent category name
                column_e = row[4] if len(row) > 4 else None
                # Column F is index 5 - for parent category name
                column_f = row[5] if len(row) > 5 else None
                # Column G is index 6 - for category name
                column_g = row[6] if len(row) > 6 else None
                
                # Skip if column G is empty
                if not column_g:
                    continue
                
                # Find grandparent category (column E) if exists
                grandparent_category = None
                if column_e:
                    grandparent_category = self.env['product.category'].search([
                        ('name', '=', column_e)
                    ], limit=1)
                
                # Find parent category (column F) if exists
                parent_category = None
                if column_f:
                    if grandparent_category:
                        # Find parent with specific grandparent
                        parent_category = self.env['product.category'].search([
                            ('name', '=', column_f),
                            ('parent_id', '=', grandparent_category.id),
                        ], limit=1)
                    else:
                        # Find parent without grandparent requirement
                        parent_category = self.env['product.category'].search([
                            ('name', '=', column_f)
                        ], limit=1)
                
                # Check if category already exists with same name, parent, and grandparent
                existing_category = None
                if parent_category:
                    existing_category = self.env['product.category'].search([
                        ('name', '=', column_g),
                        ('parent_id', '=', parent_category.id),
                    ], limit=1)
                else:
                    existing_category = self.env['product.category'].search([
                        ('name', '=', column_g)
                    ], limit=1)
                
                if existing_category:
                    skipped_count += 1
                    continue
                
                # Find account.analytic.account by code (column C)
                analytic_account = None
                if column_c:
                    analytic_account = self.env['account.analytic.account'].search([
                        ('code', '=', column_c)
                    ], limit=1)
                
                # Create new category with name from column G
                self.env['product.category'].create({
                    'name': column_g,
                    'parent_id': parent_category.id if parent_category else None,
                    'x_studio_many2one_field_2o6_1j1dfj1v3': analytic_account.id if analytic_account else None,
                    'property_valuation': 'real_time',
                    'property_cost_method': 'average',
                })
                created_count += 1
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Success'),
                    'message': _('Created %s categories, skipped %s existing') % (created_count, skipped_count),
                    'type': 'success',
                    'sticky': False,
                }
            }
            
        except Exception as e:
            raise UserError(_('Error processing Excel file: %s') % str(e))
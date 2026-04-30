import base64
import io
import logging
from collections import defaultdict

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

SHEET_NAME = 'Working File'

# 0-based column indices
COL_CODE = 1          # B - კოდი → x_studio_code
COL_BUDGET_CODE = 5   # F - ბიუჯეტის კოდი → budget_name_main.code
COL_P = 15            # P - შესაძენი
COL_Q = 16            # Q - გასაცემი
COL_R = 17            # R - საწყობის ღირებულება
COL_V = 21            # V - სულ შესაძენი
COL_W = 22            # W - სულ საწყობი
COL_X = 23            # X - სულ საწყობი+შესაძენი

# Written via ORM (regular stored fields)
FIELD_MAP_ORM = [
    (COL_P, 'x_studio_float_field_38e_1j1ftullr'),
    (COL_Q, 'x_studio_float_field_8ss_1j1ftvflr'),
    (COL_R, 'x_studio_cost_inventory'),
]

# Written via SQL (Studio computed/stored fields — ORM ignores writes to computed fields)
FIELD_MAP_SQL = [
    (COL_V, 'x_studio_float_field_tg_1j1fvkhrg'),
    (COL_W, 'x_studio_float_field_4i8_1j1fvmk7v'),
    (COL_X, 'x_studio_float_field_4nk_1j1fvng8q'),
]

FIELD_MAP = FIELD_MAP_ORM + FIELD_MAP_SQL

TARGET_REQUEST_ID = 28


class BudgetExcelImportWizard(models.TransientModel):
    _name = 'budget.excel.import.wizard'
    _description = 'Budget Excel Import Wizard'

    excel_file = fields.Binary(string='Excel File', required=True)
    file_name = fields.Char(string='File Name')
    result_message = fields.Text(string='Result', readonly=True)
    state = fields.Selection([('draft', 'Draft'), ('done', 'Done')], default='draft')

    def action_import(self):
        self.ensure_one()
        translate = _
        try:
            import openpyxl
        except ImportError:
            raise UserError(translate("openpyxl is required. Install it with: pip install openpyxl"))

        file_data = base64.b64decode(self.excel_file)
        wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)

        if SHEET_NAME not in wb.sheetnames:
            raise UserError(translate("Sheet '%s' not found in the uploaded Excel file.") % SHEET_NAME)

        ws = wb[SHEET_NAME]

        # Find header row (col B = 'კოდი', col F = 'ბიუჯეტის კოდი')
        # Limit max_col to X (col 24) to avoid MemoryError on wide sheets
        MAX_COL = COL_X + 2
        header_row_idx = None
        for row in ws.iter_rows(max_col=MAX_COL):
            b_val = row[COL_CODE].value if len(row) > COL_CODE else None
            f_val = row[COL_BUDGET_CODE].value if len(row) > COL_BUDGET_CODE else None
            if b_val and 'კოდი' in str(b_val) and f_val and 'ბიუჯეტის' in str(f_val):
                header_row_idx = row[0].row
                break

        if not header_row_idx:
            raise UserError(translate(
                "Could not find header row (B='კოდი', F='ბიუჯეტის კოდი') in sheet '%s'.") % SHEET_NAME)

        _logger.info("BudgetExcelImportWizard: header row=%s, processing data rows...", header_row_idx)

        # Aggregate values per (code, budget_code) — sum numeric fields
        data = defaultdict(lambda: {col: 0.0 for col, _f in FIELD_MAP})

        for row in ws.iter_rows(min_row=header_row_idx + 1, max_col=MAX_COL):
            if len(row) <= max(COL_CODE, COL_BUDGET_CODE):
                continue
            code = row[COL_CODE].value
            budget_code = row[COL_BUDGET_CODE].value
            if not code or not budget_code:
                continue
            code = str(code).strip()
            budget_code = str(budget_code).strip()
            key = (code, budget_code)
            for col_idx, _fname in FIELD_MAP:
                if col_idx < len(row):
                    val = row[col_idx].value
                    if val is not None:
                        try:
                            data[key][col_idx] += float(val)
                        except (TypeError, ValueError):
                            pass

        _logger.info("BudgetExcelImportWizard: aggregated %d unique (code, budget_code) groups", len(data))
        for (c, bc), fv in list(data.items())[:10]:
            _logger.info("  sample group: code=%s budget_code=%s vals=%s", c, bc,
                         {col: fv[col] for col, _ in FIELD_MAP})

        if not data:
            self.write({
                'result_message': translate("No data rows found in sheet '%s'.") % SHEET_NAME,
                'state': 'done',
            })
            return self._reload()

        request = self.env['budgeting.request'].browse(TARGET_REQUEST_ID)
        if not request.exists():
            raise UserError(translate("Budgeting request with ID %d not found.") % TARGET_REQUEST_ID)

        # Log ALL budgeting.line records for request 28 to see what x_studio_code and budget_name_main.code look like
        all_lines = self.env['budgeting.line'].search([('request_id', '=', TARGET_REQUEST_ID)])
        _logger.info("BudgetExcelImportWizard: request %d has %d budgeting lines total", TARGET_REQUEST_ID, len(all_lines))
        for bl in all_lines[:20]:
            _logger.info("  line id=%s x_studio_code=%s budget_name_main=%s (code=%s)",
                         bl.id,
                         getattr(bl, 'x_studio_code', 'FIELD_MISSING'),
                         bl.budget_name_main.name if bl.budget_name_main else None,
                         bl.budget_name_main.code if bl.budget_name_main else None)

        updated = 0
        not_found = []

        for (code, budget_code), field_sums in data.items():
            # Match account.analytic.account by code field (budget_name_main.code)
            analytic = self.env['account.analytic.account'].search(
                [('code', '=', budget_code)], limit=1)
            if not analytic:
                # Fallback: match by name
                analytic = self.env['account.analytic.account'].search(
                    [('name', '=', budget_code)], limit=1)

            _logger.info("BudgetExcelImportWizard: processing code=%s budget_code=%s → analytic id=%s name=%s code=%s",
                         code, budget_code,
                         analytic.id if analytic else None,
                         analytic.name if analytic else None,
                         analytic.code if analytic else None)

            domain = [
                ('request_id', '=', TARGET_REQUEST_ID),
                ('x_studio_code', '=', code),
            ]
            if analytic:
                domain.append(('budget_name_main', '=', analytic.id))

            lines = self.env['budgeting.line'].search(domain)
            _logger.info("  domain=%s → found %d lines", domain, len(lines))

            if not lines:
                not_found.append('%s / %s' % (code, budget_code))
                continue

            orm_vals = {fname: field_sums[col_idx] for col_idx, fname in FIELD_MAP_ORM}
            sql_vals = {fname: field_sums[col_idx] for col_idx, fname in FIELD_MAP_SQL}
            _logger.info("  orm_vals=%s  sql_vals=%s  lines=%s", orm_vals, sql_vals, lines.ids)

            for line in lines:
                try:
                    # ORM write for regular fields
                    line.write(orm_vals)
                    _logger.info("  ORM write OK for line id=%s", line.id)

                    # SQL write for Studio computed fields
                    set_clause = ', '.join('%s = %%s' % col for col in sql_vals)
                    self.env.cr.execute(
                        'UPDATE budgeting_line SET %s WHERE id = %%s' % set_clause,
                        list(sql_vals.values()) + [line.id]
                    )
                    line.invalidate_recordset()
                    _logger.info("  SQL write OK for line id=%s vals=%s", line.id, sql_vals)

                    updated += 1
                except Exception as e:
                    _logger.error("BudgetExcelImportWizard: failed to update line %s: %s", line.id, e)
                    not_found.append('%s / %s (error: %s)' % (code, budget_code, e))

        result_lines = [translate("Updated: %d line(s).") % updated]
        if not_found:
            preview = not_found[:30]
            result_lines.append(
                translate("Not matched (%d):") % len(not_found) + '\n' +
                '\n'.join('  • ' + s for s in preview)
            )
            if len(not_found) > 30:
                result_lines.append(translate("... and %d more.") % (len(not_found) - 30))

        _logger.info("BudgetExcelImportWizard: done. updated=%d not_found=%d", updated, len(not_found))
        self.write({'result_message': '\n'.join(result_lines), 'state': 'done'})
        return self._reload()

    def _reload(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

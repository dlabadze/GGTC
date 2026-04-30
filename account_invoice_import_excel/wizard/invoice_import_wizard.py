from odoo import models, fields, api, _
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
import requests
import xml.etree.ElementTree as ET
import xlwt
import io
import re
import base64
import openpyxl
from odoo.exceptions import UserError
from datetime import datetime, date
import logging

_logger = logging.getLogger(__name__)

class InvoiceImportWizard(models.TransientModel):
    _name = "invoice.import.wizard"
    _description = "Import Invoices from Excel"

    journal_id = fields.Many2one(
        'account.journal',
        string="ჟურნალი",
        required=True,
        domain=[('type', '=', 'sale')],
    )

    line_from = fields.Integer(string='პირველი ხაზი')
    line_to = fields.Integer(string='ბოლო ხაზი')
    invoice_date = fields.Date(string='ინვოისის თარიღი')

    file = fields.Binary("File", required=True)
    filename = fields.Char("Filename")
    def _validate_vat_on_rs(self, vat):
        if not vat:
            return False

        vat = str(vat).strip().replace('\xa0', '')
        if vat.endswith('.0'):
            vat = vat[:-2]

        if not re.match(r'^(\d{9}|\d{11})$', vat):
            return False

        usn = self.env.user.rs_acc
        usp = self.env.user.rs_pass
        if not usn or not usp:
            return False

        soap_request = f"""<?xml version="1.0" encoding="utf-8"?>
        <soap:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">
          <soap:Body>
            <get_name_from_tin xmlns="http://tempuri.org/">
              <su>{usn}</su><sp>{usp}</sp><tin>{vat}</tin>
            </get_name_from_tin>
          </soap:Body>
        </soap:Envelope>"""

        headers = {
            "Content-Type": "text/xml; charset=utf-8",
            "SOAPAction": "http://tempuri.org/get_name_from_tin"
        }

        try:
            response = requests.post(
                "http://services.rs.ge/waybillservice/waybillservice.asmx",
                data=soap_request,
                headers=headers,
                timeout=5
            )

            if response.status_code == 200:
                root = ET.fromstring(response.content)
                result_node = root.find('.//{http://tempuri.org/}get_name_from_tinResult')

                if result_node is None:
                    return False

                name_text = result_node.text

                if not name_text:
                    return False

                name_text = name_text.strip()
                if not name_text:
                    return False

                if name_text.lower() == 'null':
                    return False

                return True

        except Exception as e:
            _logger.error(f"RS Validation Failed for {vat}: {e}")
            return False

        return False

    def _export_error_report(self, failed_lines):
        workbook = xlwt.Workbook(encoding='utf-8')
        sheet = workbook.add_sheet('Failed Imports')
        headers = ['VAT Code', 'Partner in Excel', 'Reason']
        for i, h in enumerate(headers): sheet.write(0, i, h)
        for row_no, data in enumerate(failed_lines, start=1):
            sheet.write(row_no, 0, data['vat'])
            sheet.write(row_no, 1, data['excel_name'])
            sheet.write(row_no, 2, data['reason'])
        fp = io.BytesIO()
        workbook.save(fp)
        export_id = self.env['ir.attachment'].create({
            'name': 'failed_partners_report.xls',
            'type': 'binary',
            'datas': base64.b64encode(fp.getvalue()),
            'mimetype': 'application/vnd.ms-excel'
        })
        return {'type': 'ir.actions.act_url', 'url': f'/web/content/{export_id.id}?download=true', 'target': 'self'}


    def _sanitize_vat(self, value):
        if not value:
            return False
        val_str = str(value).strip()
        if val_str.endswith('.0'):
            val_str = val_str[:-2]
        return val_str

    def action_import_invoices(self):
        if not self.file:
            raise UserError(_("გთხოვთ ატვირთოთ ფაილი სანამ იმპორტს დაიწყებთ."))

        try:
            file_content = base64.b64decode(self.file)
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
            sheet = wb.active
        except Exception as e:
            raise UserError(("Excel ფაილის წაკითხვა ვერ მოხერხდა: %s") % str(e))

        invoices = self.env['account.move']

        def parse_number(value):
            if value is None:
                return 0.0
            try:
                return float(str(value).replace(',', '').strip())
            except Exception:
                return 0.0

        journal_name = self.journal_id.name.strip() if self.journal_id else ""
        Move = self.env['account.move']
        Partner = self.env['res.partner']  # Define Partner model shortcut
        invoice_year = self.invoice_date.year if self.invoice_date else fields.Date.today().year
        failed_lines = []

        try:
            if "ტრანსპორტირება" in journal_name:
                sheet = wb[wb.sheetnames[1]]
                invoice_counters = {}
                Product = self.env['product.product']
                ids = Product.search_read([('name', '=', 'გაზის ტრანსპორტირება')], fields=['id'], limit=1)
                product = Product.browse(ids[0]['id']) if ids else False
                if not product:
                    product = self.env['product.product'].create({'name': 'გაზის ტრანსპორტირება', 'type': 'service'})

                header_row_index = None
                for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    low_cells = [str(c).strip() if c else '' for c in row]
                    if 'Tax ID' in low_cells or 'რაოდენობა' in low_cells:
                        header_row_index = i
                        break
                if header_row_index is None:
                    header_row_index = 1

                min_row = self.line_from or 2
                max_row = self.line_to if self.line_to else sheet.max_row
                headers = [str(cell).strip() if cell else '' for cell in
                           next(sheet.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))]

                col_index = {header: idx for idx, header in enumerate(headers)}

                for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
                    partner_name = str(row[col_index.get('დასახელება RS.ge-ზე', 0)]).strip() if row[col_index.get('დასახელება RS.ge-ზე', 0)] else False

                    identification_code = self._sanitize_vat(row[col_index.get('Tax ID', 0)])
                    monthly_amount = parse_number(row[col_index.get('რაოდენობა', 0)])

                    if not monthly_amount:
                        continue

                    if not partner_name:
                        continue

                    partner = False

                    if identification_code:
                        partner = Partner.search([('vat', '=', identification_code)], limit=1)

                    if not partner:
                        if self._validate_vat_on_rs(identification_code):
                            partner = Partner.create({
                                'name': partner_name,
                                'vat': identification_code,
                                'company_type': 'company' if any(
                                    x in partner_name for x in ['შპს', 'იმ', 'სს']) else 'person'
                                })
                        else:
                            failed_lines.append({
                                'vat': identification_code,
                                'excel_name': partner.name,
                                'reason': 'Invalid VAT on RS.GE'
                            })
                            continue
                    write_vals = {}

                    if identification_code and partner.vat != identification_code:
                        write_vals['vat'] = identification_code

                    new_type = 'company' if any(
                                x in partner_name for x in ['შპს', 'იმ', 'სს']) else 'person'
                    if partner.company_type != new_type:
                        write_vals['company_type'] = new_type

                    if write_vals:
                        partner.write(write_vals)

                    prefix = 'GT'

                    invoice_name = self._get_next_invoice_name(prefix, invoice_year, Move, invoice_counters)

                    move = self.env['account.move'].create({
                        'name': invoice_name,
                        'move_type': 'out_invoice',
                        'partner_id': partner.id,
                        'invoice_date': self.invoice_date,
                        'invoice_date_due': self.invoice_date,
                        'identification_code': identification_code or '',
                        'journal_id': self.journal_id.id,
                        'invoice_line_ids': [(0, 0, {
                            'product_id': product.id,
                            'name': 'გაზის ტრანსპორტირება',
                            'quantity': monthly_amount,
                        })]
                    })
                    invoices |= move

            if failed_lines:
                return self._export_error_report(failed_lines)

            if not invoices:
                raise UserError(_("არ შეიქმნა არცერთი ინვოისი. გთხოვთ გადაამოწმოთ ფაილი."))

        except UserError:
            raise

        except Exception as e:
            import traceback
            traceback.print_exc()
            raise UserError(_("ინვოისების იმპორტისას მოხდა შეცდომა: %s") % str(e))

        return {
            "type": "ir.actions.act_window",
            "name": "Invoices",
            "res_model": "account.move",
            "view_mode": "list,form",
            "domain": [("id", "in", invoices.ids)],
            "context": {'default_move_type': 'out_invoice'}
        }

    def _get_next_invoice_name(self, prefix, year, Move, counters):
        key = (prefix, year)
        if key not in counters:
            last = Move.search([
                ('move_type', '=', 'out_invoice'),
                ('name', 'like', f'{prefix}/{year}/%')
            ], order='id desc', limit=1)
            last_index = 0
            if last and last.name:
                try:
                    last_index = int(last.name.split('/')[-1])
                except Exception:
                    pass
            counters[key] = last_index
        counters[key] += 1
        return f"{prefix}/{year}/{str(counters[key]).zfill(3)}"

    @staticmethod
    def parse_excel_date(value):
        if not value:
            return False

        if isinstance(value, (datetime, date)):
            return value.date() if isinstance(value, datetime) else value

        if isinstance(value, (int, float)):
            try:
                return from_excel(value).date()
            except Exception:
                return False

        if isinstance(value, str):
            value = " ".join(value.strip().split())
            for fmt in (
                    "%d-%m-%Y %H:%M:%S",
                    "%Y-%m-%d %H:%M:%S",
                    "%d/%m/%Y %H:%M:%S",
                    "%Y/%m/%d %H:%M:%S",
                    "%d-%m-%Y",
                    "%Y-%m-%d",
                    "%d/%m/%Y",
                    "%m/%d/%Y",
                    "%d.%m.%Y",
                    "%d.%m.%Y %H:%M:%S",
            ):
                try:
                    return datetime.strptime(value, fmt).date()
                except Exception:
                    continue

        return False
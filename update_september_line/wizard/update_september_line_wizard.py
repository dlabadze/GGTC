# -*- coding: utf-8 -*-
import base64
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import logging

_logger = logging.getLogger(__name__)


class UpdateSeptemberLineWizard(models.TransientModel):
    _name = 'update.september.line.wizard'
    _description = 'Update September Line Wizard'

    excel_file = fields.Binary(
        string='Excel File', 
        required=True, 
        help='Upload Excel file to update September Line records'
    )
    filename = fields.Char(
        string='Filename',
        help='Name of the uploaded file'
    )
    september_line_ids = fields.Many2many(
        'september.line',
        string='September Lines',
        help='September Line records to update'
    )

    def action_process_excel(self):
        """Process the uploaded Excel file"""
        self.ensure_one()
        
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file.'))
        
        try:
            # Decode the file
            file_data = base64.b64decode(self.excel_file)
            
            # Read Excel file
            excel_data = self._read_excel_file(file_data)
            
            if len(excel_data) < 2:
                raise UserError(_('Excel file must have at least a header row and one data row.'))
            
            # Process each row starting from row 1 (skip header)
            processed_count = 0
            error_lines = []
            september_lines = self.september_line_ids
            for row_idx in range(1, len(excel_data)):
                try:
                    row = excel_data[row_idx]
                    if not row or len(row) < 2:
                        continue
                    
                    # Extract code and analytic_account_id from columns
                    code = str(row[0]).strip() if row[0] else ''
                    if '.' in code:
                        code = code.split('.')[0]
                    analytic_account_id = str(row[1]).strip() if row[1] else ''
                    if code and analytic_account_id:
                        # Split analytic_account_id to extract code and name
                        analytic_code = None
                        analytic_name = None
                        
                        # Check if the format is [code] name
                        if analytic_account_id.startswith('[') and ']' in analytic_account_id:
                            # Find the closing bracket
                            bracket_end = analytic_account_id.find(']')
                            # Extract the code part (without brackets)
                            analytic_code = analytic_account_id[1:bracket_end].strip()
                            # Extract the name part (after the bracket)
                            analytic_name = analytic_account_id[bracket_end + 1:].strip()
                        else:
                            # If no brackets, treat the whole as name
                            analytic_name = analytic_account_id

                        line_with_code = september_lines.filtered(lambda l: l.x_studio_related_field_1pq_1j2ffqufh == code)
                        if line_with_code:
                            analytic_account = self.env['account.analytic.account'].search([
                                ('code', '=', analytic_code), 
                                ('name', '=', analytic_name)], limit=1)
                            if analytic_account:
                                line_with_code.budget_name_main = analytic_account.id
                            else:
                                error_lines.append(f'Row {row_idx + 1}: Analytic account not found: {analytic_account_id}')
                        else:
                            error_lines.append(f'Row {row_idx + 1}: Line with code {code} not found')
                except Exception as e:
                    error_lines.append(f'Row {row_idx + 1}: {str(e)}')
                    _logger.error(f'Error processing row {row_idx + 1}: {str(e)}')
            
            # Show result message
            message = f'Successfully processed {processed_count} row(s).'
            if error_lines:
                message += f'\n\nErrors:\n' + '\n'.join(error_lines[:10])  # Show first 10 errors
                if len(error_lines) > 10:
                    message += f'\n... and {len(error_lines) - 10} more errors'
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Processing Complete'),
                    'message': message,
                    'type': 'success' if not error_lines else 'warning',
                    'sticky': True if error_lines else False,
                }
            }
            
        except Exception as e:
            _logger.error(f'Error processing Excel file: {str(e)}')
            raise UserError(_('Error processing file: %s') % str(e))

    def _read_excel_file(self, file_data):
        """Read Excel file and return data as list of lists"""
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)
            sheet = workbook.active
            
            # Convert to list of lists
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            
            return data
            
        except ImportError:
            raise UserError(_('openpyxl library not installed. Please install: pip install openpyxl'))
        except Exception as e:
            raise UserError(_('Error reading Excel file: %s') % str(e))

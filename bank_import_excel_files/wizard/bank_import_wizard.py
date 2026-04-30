import base64  # Added missing import
import io
import openpyxl
import xlrd
from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import datetime, date
from openpyxl.utils.datetime import from_excel


class BankImportWizard(models.TransientModel):
    _name = "bank.import.wizard"
    _description = "Bank Import Excel Wizard"

    journal_id = fields.Many2one(
        'account.journal',
        string="აირჩიეთ ჟურნალი",
        required=True,
        domain=[('type', '=', 'bank')]
    )

    line_from = fields.Integer(string='პირველი ხაზი')
    line_to = fields.Integer(string='ბოლო ხაზი')

    excel_file = fields.Binary(string="აირჩიეთ Excel ფაილი", required=True)
    file_name = fields.Char(string="ფაილის სახელი")

    def action_import_excel(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError("გთხოვთ ატვირთოთ Excel ფაილი.")

        journal = self.journal_id

        journal_name_str = journal.name if journal.name else ''

        file_content = base64.b64decode(self.excel_file)
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
        sheet = wb.active

        statement_lines = self.env['account.bank.statement.line']

        def parse_number(value):
            if value is None:
                return 0.0
            try:
                s = str(value).replace('\xa0', '').replace(',', '').strip()
                if s == '':
                    return 0.0
                return float(s)
            except Exception:
                return 0.0

        if 'LB' in journal_name_str:
            sheet = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else wb.active

            header_row_index = None
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                low_cells = [str(c).strip() if c else '' for c in row]
                if 'თარიღი' in low_cells or 'Date' in low_cells:
                    header_row_index = i
                    break
            if header_row_index is None:
                header_row_index = 2

            headers = [str(c).strip() if c else '' for c in
                       next(sheet.iter_rows(min_row=header_row_index,
                                            max_row=header_row_index,
                                            values_only=True))]
            col_index = {h: i for i, h in enumerate(headers)}

            min_row = self.line_from or (header_row_index + 1)
            max_row = self.line_to or sheet.max_row

            rows = []
            for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
                if not any(row):
                    continue
                rows.append(row)

            partner_names = set()
            for row in rows:
                in_val = parse_number(row[col_index.get('შემოსული თანხა', 0)]) if 'შემოსული თანხა' in col_index else 0
                out_val = parse_number(row[col_index.get('გასული თანხა', 0)]) if 'გასული თანხა' in col_index else 0
                if in_val or out_val:
                    pname = row[col_index.get('პარტნიორი')]
                else:
                    continue
                if pname:
                    partner_names.add(str(pname).strip())

            Partner = self.env['res.partner']
            partners = Partner.search([('name', 'in', list(partner_names))])
            partner_map = {p.name: p for p in partners}
            missing_names = [n for n in partner_names if n not in partner_map]
            if missing_names:
                new_partners = Partner.create([{'name': n} for n in missing_names])
                partner_map.update({p.name: p for p in new_partners})

            line_vals_list = []

            for row in rows:
                desc = row[col_index.get('დანიშნულება', '')] or ''
                add_info = row[col_index.get('დამატებითი ინფორმაცია', '')] or ''
                payment_ref = f"{desc} / {add_info}".strip(' / ')
                date_val = row[col_index.get('თარიღი')]
                in_val = parse_number(row[col_index.get('შემოსული თანხა', 0)]) if 'შემოსული თანხა' in col_index else 0
                out_val = parse_number(row[col_index.get('გასული თანხა', 0)]) if 'გასული თანხა' in col_index else 0

                if in_val == 0 and out_val == 0:
                    continue
                if in_val:
                    amount = in_val
                else:
                    amount = -out_val
                pname = row[col_index.get('პარტნიორი')]
                ident_code = row[col_index.get('პარტნიორის საგადასახადო კოდი')]

                pname = str(pname).strip() if pname else ''

                partner = partner_map.get(pname)

                line_vals_list.append({
                    'name': payment_ref,
                    'date': date_val,
                    'partner_id': partner.id if partner else False,
                    'amount': amount,
                    'journal_id': journal.id,
                    'payment_ref': payment_ref,
                    'saidentifikacio_code': ident_code or '',
                })

            statement_lines = self.env['account.bank.statement.line'].create(line_vals_list)

        elif 'KS' in journal_name_str:
            if len(wb.sheetnames) == 1:
                sheet = wb[wb.sheetnames[0]]
            else:
                sheet = wb[wb.sheetnames[1]]
            header_row_index = None
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                low_cells = [str(c).strip() if c else '' for c in row]
                if 'თარიღი' in low_cells or 'Date' in low_cells:
                    header_row_index = i
                    break
            if header_row_index is None:
                header_row_index = 3
            headers = [str(cell).strip() if cell else '' for cell in
                       next(sheet.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))]
            col_index = {header: idx for idx, header in enumerate(headers)}

            start_row = self.line_from if self.line_from else (header_row_index + 1)
            max_row = self.line_to if self.line_to else sheet.max_row

            for row in sheet.iter_rows(min_row=start_row, max_row=max_row, values_only=True):
                if not any(cell for cell in row):
                    continue
                try :
                    desc = row[col_index.get('დანიშნულება', '')] or ''
                except Exception:
                    desc = row[col_index.get('დანიშნულება', '')] or ''

                add_info = row[col_index.get('დამატებითი ინფორმაცია', '')] or ''
                payment_ref = f"{desc}".strip(' / ')
                date_val = self.parse_excel_date(row[col_index.get('თარიღი')])
                in_val = parse_number(row[col_index.get('შემოსული თანხა', 0)]) if 'შემოსული თანხა' in col_index else 0
                out_val = parse_number(row[col_index.get('გასული თანხა', 0)]) if 'გასული თანხა' in col_index else 0
                if in_val == 0 and out_val == 0:
                    continue
                if in_val:
                    amount = in_val

                else:
                    amount = -out_val
                identification_code = row[col_index.get('საბუთის №', '')]
                line = self.env['account.bank.statement.line'].create({
                    'name': payment_ref,
                    'date': date_val,
                    'amount': amount,
                    'journal_id': journal.id,
                    'payment_ref': payment_ref,
                    'saidentifikacio_code': identification_code or '',
                })
                statement_lines |= line

        elif 'BS' in journal_name_str:
            if len(wb.sheetnames) == 1:
                sheet = wb[wb.sheetnames[0]]
            else:
                sheet = wb[wb.sheetnames[1]]
            header_row_index = None
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                low_cells = [str(c).strip() if c else '' for c in row]
                if 'თარიღი' in low_cells or 'Date' in low_cells:
                    header_row_index = i
                    break
            if header_row_index is None:
                header_row_index = 1
            headers = [str(cell).strip() if cell else '' for cell in
                       next(sheet.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))]
            col_index = {header: idx for idx, header in enumerate(headers)}

            start_row = self.line_from if self.line_from else (header_row_index + 1)
            max_row = self.line_to if self.line_to else sheet.max_row

            for row in sheet.iter_rows(min_row=start_row, max_row=max_row, values_only=True):
                if not any(cell for cell in row):
                    continue
                desc = row[col_index.get('დანიშნულება', '')] or ''
                payment_ref = f"{desc}".strip(' / ')
                date_val = self.parse_excel_date(row[col_index.get('თარიღი')])
                in_val = parse_number(row[col_index.get('ბრუნვა(კრედ)', 0)]) if 'ბრუნვა(კრედ)' in col_index else 0
                out_val = parse_number(row[col_index.get('ბრუნვა(დებ)', 0)]) if 'ბრუნვა(დებ)' in col_index else 0
                if in_val == 0 and out_val == 0:
                    continue
                if in_val:
                    amount = in_val

                else:
                    amount = -out_val
                identification_code = row[col_index.get('საბ.#', '')]
                line = self.env['account.bank.statement.line'].create({
                    'name': payment_ref,
                    'date': date_val,
                    'amount': amount,
                    'journal_id': journal.id,
                    'payment_ref': payment_ref,
                    'saidentifikacio_code': identification_code or '',
                })
                statement_lines |= line

        else:
            raise UserError(
                f"ვერ მოხერხდა ბანკის ფორმატის ამოცნობა ჟურნალის სახელიდან: {journal_name_str}. სახელი უნდა შეიცავდეს 'LB', 'KS', ან 'BS'.")

        if not statement_lines:
            raise UserError(_("არ შეიქმნა არცერთი ჩანაწერი. გთხოვთ გადაამოწმოთ ფაილი."))

        return {
            "type": "ir.actions.act_window",
            "name": "Bank Statement Lines",
            "res_model": "account.bank.statement.line",
            "view_mode": "list,form",
            "domain": [("id", "in", statement_lines.ids)],
        }

    @staticmethod
    def parse_excel_date(value):

        if not value:
            return False

        if isinstance(value, (datetime, date)):
            return value.date() if isinstance(value, datetime) else value

        if isinstance(value, (int, float)):
            try:
                return from_excel(value).date()
            except Exception:
                return False

        if isinstance(value, str):
            value = " ".join(value.strip().split())
            for fmt in (
                    "%Y/%m/%d",
                    "%d-%m-%Y %H:%M:%S",
                    "%Y-%m-%d %H:%M:%S",
                    "%d/%m/%Y %H:%M:%S",
                    "%Y/%m/%d %H:%M:%S",
                    "%d-%m-%Y",
                    "%Y-%m-%d",
                    "%d/%m/%Y",
                    "%m/%d/%Y",
                    "%d.%m.%Y",
                    "%d.%m.%Y %H:%M:%S",
            ):
                try:
                    return datetime.strptime(value, fmt).date()
                except Exception:
                    continue

        return False
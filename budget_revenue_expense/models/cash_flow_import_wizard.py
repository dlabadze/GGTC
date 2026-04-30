import base64
import io
import logging
from datetime import date

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

GREEN_FILL = 'FF00B050'
RED_FILL = 'FFFF0000'

MONTH_COLUMNS = [
    (3, 'იანვარი', date(2026, 1, 31)),
    (4, 'თებერვალი', date(2026, 2, 28)),
]


class CashFlowImportWizard(models.TransientModel):
    _name = 'cash.flow.import.wizard'
    _description = 'Cash Flow Excel Import Wizard'

    journal_id = fields.Many2one('account.journal', string='Journal', required=True)
    currency_id = fields.Many2one(
        'res.currency',
        string='Currency',
        default=lambda self: self.env.company.currency_id,
        required=True,
    )
    excel_file = fields.Binary(string='Excel File', required=True)
    file_name = fields.Char(string='File Name')
    result_message = fields.Text(string='Result', readonly=True)
    state = fields.Selection([('draft', 'Draft'), ('done', 'Done')], default='draft')

    def _get_or_create_method_line(self, payment_type):
        MethodLine = self.env['account.payment.method.line']
        lines = (self.journal_id.inbound_payment_method_line_ids
                 if payment_type == 'inbound'
                 else self.journal_id.outbound_payment_method_line_ids)

        line = lines.filtered(lambda m: m.payment_method_id.code == 'manual')[:1]
        if line:
            return line
        if lines:
            return lines[:1]

        manual_method = self.env['account.payment.method'].search([
            ('code', '=', 'manual'),
            ('payment_type', '=', payment_type),
        ], limit=1)
        if not manual_method:
            return MethodLine

        line = MethodLine.search([
            ('journal_id', '=', self.journal_id.id),
            ('payment_method_id', '=', manual_method.id),
        ], limit=1)
        if not line:
            line = MethodLine.create({
                'journal_id': self.journal_id.id,
                'payment_method_id': manual_method.id,
            })
        return line

    def _find_budget_analytic(self, budget_type):
        return self.env['budget.analytic'].search([
            ('budget_type', '=', budget_type),
        ], limit=1)

    def _find_budget_line(self, budget_analytic, excel_code, description=None):
        if not budget_analytic:
            return self.env['budget.line']

        # 1. Match by account_id.code (expense lines)
        line = self.env['budget.line'].search([
            ('budget_analytic_id', '=', budget_analytic.id),
            ('account_id.code', '=', excel_code),
        ], limit=1)
        if line:
            return line

        # 2. Match by x_plan2_id.code (revenue lines)
        line = self.env['budget.line'].search([
            ('budget_analytic_id', '=', budget_analytic.id),
            ('x_plan2_id.code', '=', excel_code),
        ], limit=1)
        if line:
            return line

        # 3. Match by description against account_id.name or x_plan2_id.name
        if description:
            line = self.env['budget.line'].search([
                ('budget_analytic_id', '=', budget_analytic.id),
                ('account_id.name', '=', description),
            ], limit=1)
            if line:
                return line
            line = self.env['budget.line'].search([
                ('budget_analytic_id', '=', budget_analytic.id),
                ('x_plan2_id.name', '=', description),
            ], limit=1)
            if line:
                return line

        return self.env['budget.line']

    def action_import(self):
        self.ensure_one()
        translate = _  # capture before any loop variable can shadow _
        _logger.info("=== CashFlowImportWizard: action_import START ===")
        try:
            import openpyxl
        except ImportError:
            raise UserError(translate("openpyxl library is required. Install it with: pip install openpyxl"))

        revenue_analytic = self._find_budget_analytic('revenue')
        expense_analytic = self._find_budget_analytic('expense')

        _logger.info("Revenue analytic: %s (id=%s)", revenue_analytic.name if revenue_analytic else None, revenue_analytic.id if revenue_analytic else None)
        _logger.info("Expense analytic: %s (id=%s)", expense_analytic.name if expense_analytic else None, expense_analytic.id if expense_analytic else None)

        if not revenue_analytic:
            raise UserError(translate("No revenue budget (budget_type='revenue') found in the system."))
        if not expense_analytic:
            raise UserError(translate("No expense budget (budget_type='expense') found in the system."))

        # Log all budget lines available for matching
        rev_lines = self.env['budget.line'].search([('budget_analytic_id', '=', revenue_analytic.id)])
        exp_lines = self.env['budget.line'].search([('budget_analytic_id', '=', expense_analytic.id)])
        _logger.info("Revenue budget has %d lines. account_id.codes: %s", len(rev_lines), [l.account_id.code for l in rev_lines])
        _logger.info("Revenue budget lines. x_plan2_id.codes: %s", [l.x_plan2_id.code if hasattr(l, 'x_plan2_id') else 'N/A' for l in rev_lines])
        _logger.info("Revenue budget lines. x_plan2_id.names: %s", [l.x_plan2_id.name if hasattr(l, 'x_plan2_id') else 'N/A' for l in rev_lines])
        _logger.info("Expense budget has %d lines. account_id.codes: %s", len(exp_lines), [l.account_id.code for l in exp_lines])

        file_data = base64.b64decode(self.excel_file)
        wb = openpyxl.load_workbook(io.BytesIO(file_data))
        ws = wb.active
        _logger.info("Excel loaded. Active sheet: %s, max_row=%s, max_col=%s", ws.title, ws.max_row, ws.max_column)

        # month_data[col_idx]['revenue'|'expense'] = [(budget_line, amount), ...]
        month_data = {col: {'revenue': [], 'expense': []} for col, _lbl, _dt in MONTH_COLUMNS}
        not_found = []

        for row in ws.iter_rows():
            code_cell = row[0]
            if not code_cell.value:
                continue
            fill = code_cell.fill
            if not fill or fill.fill_type != 'solid':
                continue
            if fill.fgColor.type != 'rgb':
                _logger.info("Row %s: fill type is '%s' (not rgb), skipping. value=%s", code_cell.row, fill.fgColor.type, code_cell.value)
                continue
            rgb = fill.fgColor.rgb
            if rgb not in (GREEN_FILL, RED_FILL):
                _logger.info("Row %s: rgb=%s not green/red, skipping. value=%s", code_cell.row, rgb, code_cell.value)
                continue

            # Excel uses "1/1.1" format; system uses "1/1/1" — normalize dots to slashes
            raw_code = str(code_cell.value).strip()
            excel_code = raw_code.replace('.', '/')
            is_revenue = (rgb == GREEN_FILL)
            analytic = revenue_analytic if is_revenue else expense_analytic
            bucket = 'revenue' if is_revenue else 'expense'

            description = str(row[1].value).strip() if len(row) > 1 and row[1].value else None
            _logger.info("Row %s: raw_code='%s' → normalized='%s', description='%s', type=%s, rgb=%s", code_cell.row, raw_code, excel_code, description, bucket, rgb)

            budget_line = self._find_budget_line(analytic, excel_code, description=description)
            if not budget_line:
                _logger.info("Row %s: NO budget line found for code='%s' / description='%s' in %s analytic (id=%s)", code_cell.row, excel_code, description, bucket, analytic.id)
                not_found.append('%s (%s)' % (excel_code, bucket))
                continue

            _logger.info("Row %s: MATCHED budget line id=%s, account=%s", code_cell.row, budget_line.id, budget_line.account_id.code)

            for col_idx, _label, _pdate in MONTH_COLUMNS:
                if col_idx >= len(row):
                    continue
                cell_val = row[col_idx].value
                if not cell_val:
                    continue
                amount = float(cell_val)
                if amount == 0:
                    continue
                _logger.info("  col=%s amount=%s → adding to %s", col_idx, abs(amount), bucket)
                month_data[col_idx][bucket].append((budget_line, abs(amount)))

        _logger.info("=== month_data summary ===")
        for col_idx, month_label, _pdate in MONTH_COLUMNS:
            _logger.info("  %s: revenue=%d entries, expense=%d entries", month_label, len(month_data[col_idx]['revenue']), len(month_data[col_idx]['expense']))

        Payment = self.env['account.payment']
        inbound_method = self._get_or_create_method_line('inbound')
        outbound_method = self._get_or_create_method_line('outbound')
        _logger.info("inbound_method: %s, outbound_method: %s", inbound_method, outbound_method)
        created_payments = []

        for col_idx, month_label, payment_date in MONTH_COLUMNS:
            revenue_entries = month_data[col_idx]['revenue']
            expense_entries = month_data[col_idx]['expense']

            if revenue_entries:
                total = sum(amt for _bl, amt in revenue_entries)
                vals = {
                    'date': payment_date,
                    'payment_type': 'inbound',
                    'journal_id': self.journal_id.id,
                    'amount': total,
                    'budget_analytic_id': revenue_analytic.id,
                }
                if inbound_method:
                    vals['payment_method_line_id'] = inbound_method.id
                payment = Payment.create(vals)
                _logger.info("Created revenue payment id=%s, now creating %d lines", payment.id, len(revenue_entries))

                PaymentLine = self.env['account.payment.line']
                for budget_line, amt in revenue_entries:
                    existing_total = sum(PaymentLine.search([('budget_line_id', '=', budget_line.id)]).mapped('amount'))
                    _logger.info("  Revenue line: budget_line=%s code=%s amount=%s existing_total=%s budget_amount=%s",
                                 budget_line.id, budget_line.account_id.code, amt, existing_total, budget_line.budget_amount)
                    try:
                        pl = PaymentLine.create({
                            'payment_id': payment.id,
                            'budget_line_id': budget_line.id,
                            'budget_analytic_id': revenue_analytic.id,
                            'currency_id': self.currency_id.id,
                            'amount': amt,
                        })
                        _logger.info("  Created payment line id=%s", pl.id)
                    except Exception as e:
                        _logger.info("  FAILED budget_line=%s code=%s amount=%s existing_total=%s budget_amount=%s error=%s",
                                     budget_line.id, budget_line.account_id.code, amt, existing_total, budget_line.budget_amount, e)

                # Update paid_amount_revenue from all payment lines for each budget line
                for budget_line, _amt in revenue_entries:
                    all_lines = PaymentLine.search([('budget_line_id', '=', budget_line.id)])
                    budget_line.write({'paid_amount_revenue': sum(all_lines.mapped('amount'))})

                created_payments.append('%s შემოსავალი — #%s (%s lines)' % (month_label, payment.id, len(revenue_entries)))

            if expense_entries:
                total = sum(amt for _bl, amt in expense_entries)
                exp_vals = {
                    'date': payment_date,
                    'payment_type': 'outbound',
                    'journal_id': self.journal_id.id,
                    'amount': total,
                    'budget_analytic_id': expense_analytic.id,
                }
                if outbound_method:
                    exp_vals['payment_method_line_id'] = outbound_method.id
                exp_payment = Payment.create(exp_vals)
                _logger.info("Created expense payment id=%s, now creating %d lines", exp_payment.id, len(expense_entries))

                PaymentLine = self.env['account.payment.line']
                for budget_line, amt in expense_entries:
                    existing_total = sum(PaymentLine.search([('budget_line_id', '=', budget_line.id)]).mapped('amount'))
                    _logger.info("  Expense line: budget_line=%s code=%s amount=%s existing_total=%s budget_amount=%s",
                                 budget_line.id, budget_line.account_id.code, amt, existing_total, budget_line.budget_amount)
                    try:
                        pl = PaymentLine.create({
                            'payment_id': exp_payment.id,
                            'budget_line_id': budget_line.id,
                            'budget_analytic_id': expense_analytic.id,
                            'currency_id': self.currency_id.id,
                            'amount': amt,
                        })
                        _logger.info("  Created payment line id=%s", pl.id)
                    except Exception as e:
                        _logger.info("  FAILED budget_line=%s code=%s amount=%s existing_total=%s budget_amount=%s error=%s",
                                     budget_line.id, budget_line.account_id.code, amt, existing_total, budget_line.budget_amount, e)

                created_payments.append('%s ხარჯი — #%s (%s lines)' % (month_label, exp_payment.id, len(expense_entries)))

        result_lines = []
        if created_payments:
            result_lines.append(translate("Created payments:") + '\n' + '\n'.join('  • ' + p for p in created_payments))
        if not_found:
            result_lines.append(translate("Budget lines not found for:") + '\n' + '\n'.join('  • ' + c for c in not_found))
        if not result_lines:
            result_lines.append(translate("No data found. Check that account_id.code on budget lines matches Excel column A codes."))

        _logger.info("=== CashFlowImportWizard: action_import END. created=%s, not_found=%s ===", len(created_payments), len(not_found))
        self.write({'result_message': '\n\n'.join(result_lines), 'state': 'done'})
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

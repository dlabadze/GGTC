# -*- coding: utf-8 -*-
import base64
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import datetime
import logging

_logger = logging.getLogger(__name__)


class MobileDebtImportWizard(models.TransientModel):
    _name = 'mobile.debt.import.wizard'
    _description = 'Mobile Debt Import Wizard'

    excel_file = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='Filename')
    starting_row = fields.Integer(string='საწყისი სტრიქონი', default=1, required=True)
    finish_row = fields.Integer(string='საბოლოო სტრიქონი', required=True)
    start_date = fields.Date(string='Start Date', required=True)
    end_date = fields.Date(string='End Date', required=True)
    comment = fields.Char(string='Comment')

    def action_import(self):
        """Import mobile debt from Excel file"""
        self.ensure_one()
        
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file.'))
        
        if self.starting_row < 1:
            raise UserError(_('Starting row must be at least 1.'))
        
        if self.finish_row < self.starting_row:
            raise UserError(_('Finish row must be greater than or equal to starting row.'))
        
        try:
            # Decode the file
            file_data = base64.b64decode(self.excel_file)
            
            # Detect file format and read accordingly
            filename_lower = (self.filename or '').lower()
            
            if filename_lower.endswith('.xlsx'):
                sheet_data = self._read_xlsx_file(file_data)
            elif filename_lower.endswith('.xls'):
                sheet_data = self._read_xls_file(file_data)
            else:
                raise UserError(_('Unsupported file format. Please upload .xls or .xlsx file.'))
            
            if len(sheet_data) < self.finish_row:
                raise UserError(_('Excel file does not have enough rows. Finish row exceeds file size.'))
            
            # Read data from starting_row to finish_row (convert to 0-based index)
            # Expected columns (0-based index):
            # 0: First column (not used)
            # 1: Second column (not used)
            # 2: Phone number (third column)
            # 3: Debt (fourth column)
            # 4: Debt (fifth column) - if exists, may be used
            
            # Group rows by employee phone number
            employee_rows = {}  # {employee_id: [rows]}
            employee_phones = {}  # {employee_id: set(phone_str)} - collect unique phone numbers from Excel
            error_lines = []
            not_found_phones = set()  # Collect unique phone numbers that don't match any employee
            
            for row_idx in range(self.starting_row - 1, self.finish_row):  # Convert to 0-based
                if row_idx >= len(sheet_data):
                    break
                
                try:
                    row = sheet_data[row_idx]
                    if not row or len(row) == 0:
                        continue
                    
                    # Column 2 (third column): Phone number
                    phone_str = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                    # Remove decimal point if exists
                    if '.' in phone_str:
                        phone_str = phone_str.split('.')[0]
                    
                    if not phone_str:
                        error_lines.append(f'Row {row_idx + 1}: Missing phone number')
                        continue
                    
                    # Find employee by phone number
                    employee = self._find_employee_by_phone(phone_str)
                    
                    if not employee:
                        not_found_phones.add(phone_str)
                        error_lines.append(f'Row {row_idx + 1}: Employee not found with phone: {phone_str}')
                        continue
                    
                    # Get debt from fourth and fifth columns (indices 3 and 4)
                    # Convert negative values to positive (abs)
                    debt = 0.0
                    if len(row) > 3 and row[3]:
                        try:
                            debt += abs(float(row[3]))
                        except:
                            pass

                    # Group by employee
                    if employee.id not in employee_rows:
                        employee_rows[employee.id] = []
                        employee_phones[employee.id] = set()
                    
                    employee_rows[employee.id].append({
                        'employee': employee,
                        'debt': debt,
                        'row_idx': row_idx + 1
                    })
                    
                    # Collect phone number from Excel
                    employee_phones[employee.id].add(phone_str)
                    
                except Exception as e:
                    error_lines.append(f'Row {row_idx + 1}: {str(e)}')
                    _logger.error(f'Error processing row {row_idx + 1}: {str(e)}')
            
            # Create mobile.debt record
            mobile_debt = self.env['mobile.debt'].create({
                'start_date': self.start_date,
                'end_date': self.end_date,
                'comment': self.comment,
            })
            
            # Create mobile.debt.detail records - one per employee with summed debt
            created_count = 0
            for employee_id, rows in employee_rows.items():
                # Sum up all debts for this employee
                total_debt = sum(row_data['debt'] for row_data in rows)
                
                # Get the first employee record (they're all the same employee)
                employee = rows[0]['employee']
                
                # Get the phone number from employee's private_phone field
                private_phone = employee.private_phone or ''

                # Get phone numbers from Excel and join with '; ' if multiple.
                # Also normalize each number by stripping extra spaces.
                phones_from_excel = employee_phones.get(employee_id, set())
                if phones_from_excel:
                    normalized_phones = sorted(
                        {p.strip() for p in phones_from_excel if p and str(p).strip()}
                    )
                    private_phone_from_excel = "; ".join(normalized_phones)
                else:
                    private_phone_from_excel = ''
                
                # Create one detail record per employee
                self.env['mobile.debt.detail'].create({
                    'mobile_debt_id': mobile_debt.id,
                    'employee_id': employee.id,
                    'private_phone': private_phone,
                    'private_phone_from_excel': private_phone_from_excel,
                    'debt': total_debt,
                })
                created_count += 1
            
            # Build notification message
            message = f'Created {created_count} Mobile Debt Detail record(s) for {len(employee_rows)} employee(s).'
            
            # Generate and download Excel file for missed phone numbers
            if not_found_phones:
                try:
                    excel_data = self._generate_missed_phones_excel(not_found_phones)
                    
                    # Create attachment for download
                    attachment = self.env['ir.attachment'].create({
                        'name': 'missed_phones.xlsx',
                        'type': 'binary',
                        'datas': excel_data,
                        'res_model': 'mobile.debt.import.wizard',
                        'res_id': self.id,
                        'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    })
                    
                    # Return download action
                    return {
                        'type': 'ir.actions.act_url',
                        'url': f'/web/content/{attachment.id}?download=true',
                        'target': 'new',
                    }
                    
                except Exception as e:
                    _logger.error(f'Error generating Excel file: {str(e)}')
                    # Fallback to showing in message
                    phones_list = ', '.join(sorted(not_found_phones))
                    message += f'\n\n[{phones_list}] "ამ ნომრებზე თანამშრომელი ვერ მოიძებნა"'
            
            # Return notification for successful import without missed phones
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import Complete'),
                    'message': message,
                    'type': 'success',
                }
            }
            
        except Exception as e:
            _logger.error(f'Error importing mobile debt: {str(e)}')
            raise UserError(_('Error importing file: %s') % str(e))
    
    def _generate_missed_phones_excel(self, phone_numbers):
        """Generate Excel file with missed phone numbers"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Missed Phones'
            
            # Add header with Georgian text
            ws['A1'] = 'მობილურის ნომრები'
            ws['A1'].font = Font(bold=True, size=12)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Set column width
            ws.column_dimensions['A'].width = 25
            
            # Add phone numbers (sorted)
            for idx, phone in enumerate(sorted(phone_numbers), start=2):
                ws[f'A{idx}'] = phone
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            excel_data = output.getvalue()
            output.close()
            
            # Encode to base64
            return base64.b64encode(excel_data)
            
        except ImportError:
            raise UserError(_('openpyxl library not installed. Please install: pip install openpyxl'))
        except Exception as e:
            _logger.error(f'Error creating Excel file: {str(e)}')
            raise UserError(_('Error generating Excel file: %s') % str(e))
    
    def _find_employee_by_phone(self, phone_str):
        """Find employee by phone number"""
        # Search for employees where private_phone matches or contains the phone
        # First try exact match
        employee = self.env['hr.employee'].search([
            ('private_phone', '=', phone_str)
        ], limit=1)
        
        if employee:
            return employee
        
        # Try to find where phone is in a list separated by "; "
        employees = self.env['hr.employee'].search([
            ('private_phone', '!=', False),
            ('private_phone', '!=', '')
        ])
        
        for emp in employees:
            if not emp.private_phone:
                continue
            
            # Split by "; " if it exists
            if "; " in emp.private_phone:
                phones = [p.strip() for p in emp.private_phone.split("; ")]
                if phone_str in phones:
                    return emp
            else:
                # Direct match
                if emp.private_phone.strip() == phone_str:
                    return emp
        
        return False
    
    def _read_xlsx_file(self, file_data):
        """Read .xlsx file using openpyxl"""
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)
            sheet = workbook.active
            
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            
            return data
            
        except ImportError:
            raise UserError(_('openpyxl library not installed. Please install: pip install openpyxl'))
        except Exception as e:
            raise UserError(_('Error reading .xlsx file: %s') % str(e))
    
    def _read_xls_file(self, file_data):
        """Read .xls file using xlrd"""
        try:
            import xlrd
            
            workbook = xlrd.open_workbook(file_contents=file_data)
            sheet = workbook.sheet_by_index(0)
            
            data = []
            for row_idx in range(sheet.nrows):
                row = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    
                    if sheet.cell_type(row_idx, col_idx) == 3:  # XL_CELL_DATE
                        try:
                            cell_value = xlrd.xldate_as_datetime(cell_value, workbook.datemode)
                        except:
                            pass
                    
                    row.append(cell_value)
                data.append(row)
            
            return data
            
        except ImportError:
            raise UserError(_('xlrd library not installed. Please install: pip install xlrd'))
        except Exception as e:
            raise UserError(_('Error reading .xls file: %s') % str(e))


import base64
from io import BytesIO

from odoo import api, fields, models, _
from odoo.exceptions import UserError


try:
    import xlrd  # for .xls
except ImportError:  # pragma: no cover
    xlrd = None

try:
    from openpyxl import load_workbook  # for .xlsx
except ImportError:  # pragma: no cover
    load_workbook = None


class SawvysiNashtiWizard(models.TransientModel):
    _name = "sawyisi.nashti.wizard"
    _description = "Initial Balance Excel Import"

    file_data = fields.Binary(string="Excel File", required=True)
    file_name = fields.Char(string="Filename")
    start_row = fields.Integer(string="Start Row", default=2, help="First Excel row to read (1-based).")
    end_row = fields.Integer(string="End Row", required=True, help="Last Excel row to read (1-based).")
    journal_id = fields.Many2one(
        "account.journal",
        string="Journal",
        required=True,
        domain=[("type", "=", "general")],
        default=lambda self: self.env["account.journal"].search(
            [("name", "=", "Opening Entries Journal")], limit=1
        ),
    )
    date = fields.Date(
        string="Date",
        required=True,
        default=lambda self: fields.Date.to_date("2025-12-31"),
    )

    def action_import_excel(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError(_("Please upload an Excel file."))
        if self.start_row <= 0 or self.end_row <= 0 or self.end_row < self.start_row:
            raise UserError(_("Invalid start/end row range."))
        if not self.file_name:
            raise UserError(_("Filename is missing; please re-upload the file."))

        # Decide reader by extension
        filename = self.file_name.lower()
        decoded = base64.b64decode(self.file_data)

        use_xls = filename.endswith(".xls") and not filename.endswith(".xlsx")
        use_xlsx = filename.endswith(".xlsx")

        if use_xls:
            if not xlrd:
                raise UserError(_("Python library 'xlrd' is required to read .xls files."))
            try:
                workbook = xlrd.open_workbook(file_contents=decoded)
            except Exception as e:  # pragma: no cover
                raise UserError(_("Could not read .xls file: %s") % (e,))
            sheet = workbook.sheet_by_index(0)
            total_rows = sheet.nrows
        elif use_xlsx:
            if not load_workbook:
                raise UserError(_("Python library 'openpyxl' is required to read .xlsx files."))
            try:
                wb = load_workbook(filename=BytesIO(decoded), data_only=True)
            except Exception as e:  # pragma: no cover
                raise UserError(_("Could not read .xlsx file: %s") % (e,))
            sheet = wb.active
            total_rows = sheet.max_row
        else:
            raise UserError(_("Unsupported file format. Please upload .xls or .xlsx file."))

        move_lines_vals = []
        total_debit = 0.0
        company = self.env.company

        Account = self.env["account.account"].sudo().with_company(company)

        # Iterate rows (1-based indices for user, adapt per engine)
        for row_idx in range(self.start_row, min(self.end_row, total_rows) + 1):
            display_row = row_idx
            # Column B and D: index logic differs by engine
            if use_xls:
                xls_row = row_idx - 1
                debit_code_cell = sheet.cell(xls_row, 1)  # B
                amount_cell = sheet.cell(xls_row, 3)      # D
                debit_raw = debit_code_cell.value
                amount_raw = amount_cell.value
            else:  # xlsx / openpyxl
                # openpyxl rows/cols are 1-based: B=2, D=4
                debit_raw = sheet.cell(row=row_idx, column=2).value
                amount_raw = sheet.cell(row=row_idx, column=4).value

            debit_code = str(debit_raw).strip() if debit_raw else ""
            if not debit_code:
                # Skip empty rows
                continue

            try:
                amount = float(amount_raw or 0.0)
            except Exception:
                raise UserError(_("Invalid amount on row %s.") % (display_row,))

            if amount == 0.0:
                continue

            debit_account = Account.search([("code", "=", debit_code)], limit=1)
            if not debit_account:
                raise UserError(_("Account with code %s not found (row %s).") % (debit_code, display_row))

            move_lines_vals.append(
                (
                    0,
                    0,
                    {
                        "name": _("Initial balance row %s") % (display_row,),
                        "account_id": debit_account.id,
                        "debit": amount,
                        "credit": 0.0,
                        "company_id": company.id,
                    },
                )
            )
            total_debit += amount

        if not move_lines_vals:
            raise UserError(_("No valid rows found in the selected range."))

        credit_account = Account.search([("code", "=", "9999")], limit=1)
        if not credit_account:
            raise UserError(_("Credit account with code 9999 not found."))

        move_lines_vals.append(
            (
                0,
                0,
                {
                    "name": _("Initial balance total"),
                    "account_id": credit_account.id,
                    "debit": 0.0,
                    "credit": total_debit,
                    "company_id": company.id,
                },
            )
        )

        move_vals = {
            "date": self.date,
            "ref": _("Initial balance import"),
            "journal_id": self.journal_id.id,
            "line_ids": move_lines_vals,
            "company_id": company.id,
        }
        move = self.env["account.move"].sudo().create(move_vals)

        return {
            "type": "ir.actions.act_window",
            "res_model": "account.move",
            "res_id": move.id,
            "view_mode": "form",
            "target": "current",
        }


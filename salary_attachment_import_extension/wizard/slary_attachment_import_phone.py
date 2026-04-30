# -*- coding: utf-8 -*-
import base64
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
from datetime import datetime
import logging

_logger = logging.getLogger(__name__)


class SalaryAttachmentImportPhone(models.TransientModel):
    _name = 'salary.attachment.import.phone'
    _description = 'Salary Attachment Import Phone'

    excel_file = fields.Binary(string='Excel File', required=True, help='Upload Excel file with salary attachment data')
    filename = fields.Char(string='Filename')
    
    def action_import(self):
        """Import salary attachments from Excel file"""
        self.ensure_one()
        
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file.'))
        
        try:
            # Decode the file
            file_data = base64.b64decode(self.excel_file)
            
            # Detect file format and read accordingly
            filename_lower = (self.filename or '').lower()
            
            if filename_lower.endswith('.xlsx'):
                # Read .xlsx file with openpyxl
                sheet_data = self._read_xlsx_file(file_data)
            elif filename_lower.endswith('.xls'):
                # Read .xls file with xlrd
                sheet_data = self._read_xls_file(file_data)
            else:
                raise UserError(_('Unsupported file format. Please upload .xls or .xlsx file.'))
            
            # Get header row
            if len(sheet_data) < 2:
                raise UserError(_('Excel file must have at least a header row and one data row.'))
            
            # Read data starting from row 1 (skip header)
            # Expected columns (0-based index):
            # 0: Employee_ids (private_phone)
            # 1: description
            # 2: Other_input_type_id
            # 3: date_start
            # 4: date_end
            # 5: monthly_amount
            # 6: total_amount
            # 7: status
            
            created_count = 0
            error_lines = []
            for row_idx in range(1, len(sheet_data)):
                try:
                    row = sheet_data[row_idx]
                    if row == []:
                        continue
                    # Column 0: Employee_ids (private_phone)
                    employee_phone_str = str(row[0]).strip() if len(row) > 0 and row[0] else ''
                    # Remove decimal point and everything after it (e.g., "123.0" -> "123")
                    if '.' in employee_phone_str:
                        employee_phone_str = employee_phone_str.split('.')[0]
                    
                    if not employee_phone_str:
                        error_lines.append(f'Row {row_idx + 1}: Missing Employee_ids (private_phone)')
                        continue
                    
                    # Search for employee by private_phone
                    employee = self.env['hr.employee'].search([
                        ('private_phone', '=', employee_phone_str)
                    ], limit=1)
                    
                    if not employee:
                        error_lines.append(f'Row {row_idx + 1}: Employee not found with private_phone: {employee_phone_str}')
                        continue
                    
                    # Column 1: description
                    description = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    # Column 2: Other_input_type_id
                    other_input_type_id = None
                    if len(row) > 2 and row[2]:
                        other_input_type_name = str(row[2]).strip()
                        # Search for other input type by name
                        other_input_type = self.env['hr.payslip.input.type'].search([
                            ('name', '=', other_input_type_name)
                        ], limit=1)
                        if other_input_type:
                            other_input_type_id = other_input_type.id
                    
                    # Column 3: date_start
                    date_start = None
                    if len(row) > 3 and row[3]:
                        date_start = self._parse_date(row[3])
                    
                    # Column 4: date_end
                    date_end = None
                    if len(row) > 4 and row[4]:
                        date_end = self._parse_date(row[4])
                    
                    # Column 5: monthly_amount
                    monthly_amount = 0.0
                    if len(row) > 5 and row[5]:
                        try:
                            monthly_amount = float(row[5])
                        except:
                            monthly_amount = 0.0
                    
                    # Column 6: total_amount
                    total_amount = 0.0
                    if len(row) > 6 and row[6]:
                        try:
                            total_amount = float(row[6])
                        except:
                            total_amount = 0.0
                    
                    # Column 7: status
                    status = 'open'
                    if len(row) > 7 and row[7]:
                        status_value = str(row[7]).strip().lower()
                        if status_value in ['open', 'close', 'cancel']:
                            status = status_value
                        elif status_value in ['running', 'დასრულებული', 'გაუქმებული']:
                            if status_value == 'running':
                                status = 'open'
                            elif status_value == 'დასრულებული':
                                status = 'close'
                            elif status_value == 'გაუქმებული':
                                status = 'cancel'
                    
                    # Create salary attachment
                    vals = {
                        'employee_ids': [(6, 0, [employee.id])],
                        'description': description,
                        'monthly_amount': monthly_amount,
                        'total_amount': total_amount,
                    }
                    
                    # Add optional fields if they exist
                    if other_input_type_id:
                        vals['other_input_type_id'] = other_input_type_id
                    if date_start:
                        vals['date_start'] = date_start
                    if date_end:
                        vals['date_end'] = date_end
                    if status:
                        vals['state'] = status
                    
                    self.env['hr.salary.attachment'].create(vals)
                    created_count += 1
                    
                except Exception as e:
                    error_lines.append(f'Row {row_idx + 1}: {str(e)}')
                    _logger.error(f'Error processing row {row_idx + 1}: {str(e)}')
            
            # Show result message
            message = f'Successfully imported {created_count} salary attachment(s).'
            if error_lines:
                message += f'\n\nErrors:\n' + '\n'.join(error_lines[:10])  # Show first 10 errors
                if len(error_lines) > 10:
                    message += f'\n... and {len(error_lines) - 10} more errors'
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import Complete'),
                    'message': message,
                    'type': 'success' if not error_lines else 'warning',
                    'sticky': True if error_lines else False,
                }
            }
            
        except Exception as e:
            _logger.error(f'Error importing salary attachments: {str(e)}')
            raise UserError(_('Error importing file: %s') % str(e))

    def _read_xlsx_file(self, file_data):
        """Read .xlsx file using openpyxl"""
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)
            sheet = workbook.active
            
            # Convert to list of lists
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            
            return data
            
        except ImportError:
            raise UserError(_('openpyxl library not installed. Please install: pip install openpyxl'))
        except Exception as e:
            raise UserError(_('Error reading .xlsx file: %s') % str(e))

    def _read_xls_file(self, file_data):
        """Read .xls file using xlrd"""
        try:
            import xlrd
            
            workbook = xlrd.open_workbook(file_contents=file_data)
            sheet = workbook.sheet_by_index(0)
            
            # Convert to list of lists
            data = []
            for row_idx in range(sheet.nrows):
                row = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    
                    # Convert Excel dates
                    if sheet.cell_type(row_idx, col_idx) == 3:  # XL_CELL_DATE
                        try:
                            cell_value = xlrd.xldate_as_datetime(cell_value, workbook.datemode)
                        except:
                            pass
                    
                    row.append(cell_value)
                data.append(row)
            
            return data
            
        except ImportError:
            raise UserError(_('xlrd library not installed. Please install: pip install xlrd'))
        except Exception as e:
            raise UserError(_('Error reading .xls file: %s') % str(e))

    def _parse_date(self, date_value):
        """Parse date from various formats"""
        if not date_value:
            return None
        
        try:
            # If already a datetime object
            if isinstance(date_value, datetime):
                return date_value.date()
            
            # If it's a date object
            from datetime import date as date_type
            if isinstance(date_value, date_type):
                return date_value
            
            # If it's a string, try to parse
            if isinstance(date_value, str):
                date_str = date_value.strip()
                
                # Try different formats (with and without leading zeros)
                date_formats = [
                    '%Y-%m-%d',      # 2025-05-25
                    '%d/%m/%Y',      # 25/05/2025
                    '%m/%d/%Y',      # 05/25/2025 or 5/25/2025
                    '%Y/%m/%d',      # 2025/05/25
                    '%d-%m-%Y',      # 25-05-2025
                    '%m-%d-%Y',      # 05-25-2025
                ]
                
                for fmt in date_formats:
                    try:
                        return datetime.strptime(date_str, fmt).date()
                    except:
                        continue
                
                # If standard formats don't work, try parsing manually for M/D/Y format
                if '/' in date_str:
                    parts = date_str.split('/')
                    if len(parts) == 3:
                        try:
                            # Try M/D/Y format (5/25/2025)
                            month, day, year = int(parts[0]), int(parts[1]), int(parts[2])
                            from datetime import date as date_type
                            return date_type(year, month, day)
                        except:
                            try:
                                # Try D/M/Y format
                                day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                                from datetime import date as date_type
                                return date_type(year, month, day)
                            except:
                                pass
            
            # If it's a number (Excel serial date)
            if isinstance(date_value, (int, float)):
                # Excel date serial number
                from datetime import timedelta
                base_date = datetime(1899, 12, 30)
                return (base_date + timedelta(days=int(date_value))).date()
            
        except Exception as e:
            _logger.warning(f'Could not parse date: {date_value}, error: {str(e)}')
        
        return None
